print()
print()
# Exercise Data Converter - Converts .tcx file exercise data from Garmin device to an .xlsx format.
# Version 1.0
# Written by Tom Reid
# Python 3.8.3

#----------------------------------------------------------------------------------------------
''' NOTES '''

# Run as a #! file with a batch file later
# Use glob() to find all relevant files (.TCX) -- COMPLETE
# Check folder for unconverted files. Convert any of these.  (Use regex for this) -- COMPLETE
# Save list of all files converted to csv. -- COMPLETE
# Add some input validation maybe? To check file names? Don't know if it is necessary really because of the regex's 

# Ideally in order to ensure that each piece of data is connected, first you would find each 
# chunk of data, and from there remove the data. That way you can be sure it is coming form the same source.  -- COMPLETE

# Check that each bit of data is tied together. With a counter for time value print maybe? -- COMPLETE
# Define the regex outside of the main loop? More efficient probs   -- COMPLETE

# Add a regex to search for the type of exercise and the date for the spreadsheet -- COMPLETE

# Check that all lists in saved_data are the same length to verify printing - unnecessary?

# Centralise file location names

# Ensure excel files won't be overwritten -- COMPLETE

# Make paths relative to the main folder.

# Save the excel files in a separate folder


'''Questions'''

# Q:Why can't I use a regex which finds the heartbeat using the lines above and below the value
#   which say heart bpm or something? How am I doing it wrong?
#   Attempts:
    #heartrateRegex = re.compile(r'.*<HeartRateBpm>(.*).*</HeartRateBpm>', re.DOTALL)
    #heartrateRegex = re.compile(r'\s{12}<HeartRateBpm>\n\s{14}(.*)\n\s{12}</HeartRateBpm>')
# A:You are analysing the file line by line, which is why it is impossible to find a set of lines

#----------------------------------------------------------------------------------------------

from pathlib import Path    # Module for finding files along a path 
import os
import re   # Module for regular expressions
import openpyxl
from datetime import time
import random
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, numbers, Alignment  # why do I have to do this if I have already imported all of openpyxl?

filenamesRegex = re.compile(r'(.*).tcx')    # Create a regex for finding the files from which we want to extract data
fileslocation = Path(__file__).parent / "./"    # Gives the absolute path of the .py file
exercisefile_locations = []

# Find the any .tcx files in the specified location and add them to a list 
filelist = list(fileslocation.glob('*'))    # Puts all (including non tcx) file locations in the folder into a list
for i in filelist:
    mo_filenames = filenamesRegex.search(str(i))    # Searches for the file locations/names with the .tcx suffix
    if mo_filenames != None:                        # If a file with the .tcx suffix is found, this trigger
        exercisefile_locations.append(str(i))       # Adds the location of any .tcx files to the list exercisefile_locations 

# Check if files have been edited before by referencing list of previously edited files
converted_list_path = Path(__file__).parent / ".//1. ConvertedList.txt"
converted_files_list = [] 
try:
    with open(converted_list_path, 'rt') as convlist:    
        for lines in convlist:    # Goes through each line in the file 
            mo_filenames = filenamesRegex.search(str(lines))
            if mo_filenames != None:
                converted_files_list.append(lines.rstrip('\n'))   # Puts previously converted file names into a list

# If folder with converted files cannot be opened, it does not exist. This except creates it.  
except: 
    convertedfiles_file = open(converted_list_path, 'w')
    convertedfiles_file.write('DO NOT DELETE THIS FILE! \n\nConverted files: \n\n')
    convertedfiles_file.close()
# Honestly the variable names in this ^ section are a mess. They need fixing. 

dellist = []
dellistprint = []
# Create list of files which have not been converted
for iindex, i in enumerate(exercisefile_locations):
    for jindex, j in enumerate(converted_files_list):
        if str(i) == str(j):
            dellist.append(iindex)
            dellistprint.append(i)

convertedstatus = False # Will be used to measure whether any previously converted files are found. 
for i in sorted(dellist, reverse = True):   # Have to remove them in reverse order or it messes things up
    del exercisefile_locations[i]
    convertedstatus = True

if convertedstatus == True:    
    print('Converted tcx files found. The following files will not be converted as they have already been converted: \n')
    for i in dellistprint:
        print(i)

# Regular expression definitions
timeRegex = re.compile(r'<Time>\d\d\d\d\W\d\d\W\d\d\D(.*)\W(.*)\W(.*)\W(.*)\D</Time>')  #timeRegex = re.compile(r'<Time>\d\d\d\d\W\d\d\W\d\d\D(.*)\W(.*)\W(.*)\W(.*)\D</Time>')
heartrateRegex = re.compile(r'<Value>(.*)</Value>')     # Next time use Beautiful Soup module to handle regex's.
altitudeRegex = re.compile(r'<AltitudeMeters>(.*)</AltitudeMeters>')
distanceRegex = re.compile(r'<DistanceMeters>(.*)</DistanceMeters>')
speedRegex = re.compile(r'<ns3:Speed>(.*)</ns3:Speed>')
cadenceRegex = re.compile(r'<Cadence>(.*)</Cadence>')
exercisetypeRegex = re.compile(r'<Activity Sport="(.*)">')
dateRegex = re.compile(r'<Id>(.*)\W(.*)\W(.*)\D\d\d\D\d\d\W\d\d\W\d\d\d\D</Id>')    # <Id>2020-06-03T05:43:47.000Z</Id> -- example  ##Q: Why does this work?

# Create a folder for the excel files if it does not already exist
excel_folder_path = Path(__file__).parent / 'Excel Files'
if excel_folder_path.exists():
    print('Excel folder exists.')
else:
    os.makedirs(excel_folder_path)


# Main file loop. This extracts the relevant data from the tcx files. 
for fileindex, file in enumerate(exercisefile_locations):   # Maybe this should be a function? And then we should do a small loop which calls the function?
    currentfile_location = file # Change this each loop
    filedatadump = []
    with open(Path(currentfile_location),'rt') as exdata:   # Performs any actions needed while the file is open. Closes the file following.
        
        for lines in exdata:    # Goes through each line in the file 
            filedatadump.append(lines.rstrip('\n'))  # Puts each line into a list. rstrip removes the new line at the end of each line. 

    exercisetype = 'Unknown'
    exercisedate = []

    #t_incrementdata = ['time', 'heartrate', 'altitude', 'distance', 'speed', 'cadence'] # Relevant data will be put into this list
    t_incrementdata = [0, 0, 0, 0, 0, 0]
    saved_data = [] # Lists of relevant data will be put into this list. Collected so that data is connected to its timestamp (ideally)

    ## Search filedatadump for relevant information using regular expressions (regex)
    for i in filedatadump: # Runs through each line in the file
        # 2020-06-03T05:43:47.000Z - Example time stamp
        mo_time = timeRegex.search(i)   # mo - Matching Objects
        if mo_time != None:
            t_incrementdata = [0, 0, 0, 0, 0, 0] # Resetting input list
            t_incrementdata[0] = mo_time[1] + ':' + mo_time[2] + ':' + mo_time[3]
            #print(str(mo_time[1]) + ':' + str(mo_time[2]) + ':'+ str(mo_time[3]))
            saved_data.append(t_incrementdata)    # Adds data from previous section into list
            continue

        # Search for heartrate values
        mo_heartrate = heartrateRegex.search(i) # you don't need to do this if any of the previous regexes have been found. More efficient if the code runs slowly. 
        if mo_heartrate != None:
            t_incrementdata[1] = int(mo_heartrate[1])
            continue
        
        # Search for altitudes
        mo_altitude = altitudeRegex.search(i)
        if mo_altitude != None:
            t_incrementdata[2] = float(mo_altitude[1])
            continue
        
        # Search for distance from start values
        mo_distance = distanceRegex.search(i)
        if mo_distance != None:
            t_incrementdata[3] = float(mo_distance[1])
            continue
        
        #Search for speed values
        mo_speed = speedRegex.search(i)
        if mo_speed != None:
            t_incrementdata[4] = float(mo_speed[1])
            continue

        #Search for cadence values
        mo_cadence = cadenceRegex.search(i)
        if mo_cadence != None:
            t_incrementdata[5] = int(mo_cadence[1])
            continue
        
        # Search for exercise type
        mo_exercisetype = exercisetypeRegex.search(i)
        if mo_exercisetype != None:
            exercisetype = mo_exercisetype[1]
            continue 

        # Search for exercise date
        mo_exercisedate = dateRegex.search(i)
        if mo_exercisedate != None:
            #exercisedate = str(mo_exercisedate[1]) + '.' + str(mo_exercisedate[2]) + '.' + str(mo_exercisedate)
            exercisedate.append(str(mo_exercisedate[1]) + '.' + str(mo_exercisedate[2]) + '.' + str(mo_exercisedate[3])) #+ ' ' + str(mo_exercisedate[4]))
            continue

    ## Excel Sheet Work ## -----------------------------------------------------------------------------------------------------
    wb = Workbook()     # Creates excel sheet
    ws = wb.active    # Selects active worksheet in the excel file 
    ws.title = str(exercisetype) + ' Data' 

    # Creating a font for the title cells
    titles = NamedStyle(name = 'titles')
    titles.font = Font(bold = True, size = 11)
    titles.border = Border(bottom = Side(style = 'thin'))
    titles.number_format = numbers.FORMAT_TEXT

    # Title the columns
    ws.cell(row = 1, column = 1, value = 'Time (Absolute)')
    ws.cell(row = 1, column = 2, value = 'Heart Rate (BPM)')
    ws.cell(row = 1, column = 3, value = 'Altitude')
    ws.cell(row = 1, column = 4, value = 'Distance')
    ws.cell(row = 1, column = 5, value = 'Speed')
    ws.cell(row = 1, column = 6, value = 'Cadence')

    # Applying the font to the title cells
    wb.add_named_style(titles) # Making the style a known style to the module
    ws['A1'].style = 'titles'
    ws['B1'].style = 'titles'
    ws['C1'].style = 'titles'
    ws['D1'].style = 'titles'
    ws['E1'].style = 'titles'
    ws['F1'].style = 'titles'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20

    # Write relevant data to excel sheet
    for yindex, row in enumerate(range(3, len(saved_data) + 3)):
        for xindex, column in enumerate(range(1, len(saved_data[1]) + 1)):
            cell = ws.cell(row = row, column = column)
            cell.value = saved_data[yindex][xindex]

    # Insert and format a column for relative time values
    ws.insert_cols(2)   
    ws.cell(row = 1, column = 2, value = 'Time Since Start')
    ws['B1'].style = 'titles'
    ws.column_dimensions['B'].width = 20

    # Creating another named format to format the time columns as time
    times = NamedStyle(name = 'times')
    times.font = Font(size = 11)
    wb.add_named_style(times)
    #times.number_format = numbers.FORMAT_TIME
    times.number_format = numbers.FORMAT_DATE_TIME4

    for x, cells in enumerate(ws['A']):

        if x > 1:
            ws['A' + str(x + 1)].style = 'times'
            ws['B' + str(x + 1)].style = 'times'
        
        if x > 2:
            ws['B' + str(x + 1)] = '=(A' + str(x + 1) + '-A' + str(x) + ')' + '+B' + str(x) # Entering excel formula into all cells in B
            

    ws['B3'] = '00:00:00'
    ws['B3'].alignment = Alignment(horizontal='right', vertical='bottom')   # Aligns B3 correctly because for some reason it likes to left justify

    # Check the names of existing excel files to ensure we don't save over one.
    excelfilenamesRegex = re.compile(r'(.*).xlsx')    # Create a regex for finding the files from which we want to extract data
    #excelfiles_location = Path('C:/Users/tomja/Documents/5 - Learning/Python/DadExerciseData/Raw Data')    # Change this to be the location of the excel files
    excelfiles_location = Path(__file__).parent / './'
    excel_filelocations = []

    # Find the existing excel files
    excelfilelist = list(excelfiles_location.glob('*'))    # Puts all file locations (names) into a list
    for i in excelfilelist:
        mo_filenames = excelfilenamesRegex.search(str(i))    # Searches for the file locations with the .tcx suffix
        if mo_filenames != None:
            excel_filelocations.append(str(i))  # Saves a list of all excel files already in the folder

    # Getting the time as a string to add to the file name
    savetimestart = ''
    for counter, i in enumerate(saved_data[0][0]):  # This is a dumb way of doing it. Use split function.        
        
        try:
            (int(i))
            a = int(i)
            savetimestart = savetimestart + str(a)

            if counter >= 4:
                break

        except:
            pass

    
    potential_save_name_str = './/TemporaryName' + str(random.randint(0,1000)) + '.xlsx'
    potential_save_name = Path(__file__).parent / potential_save_name_str

    # Generate the file name
    try:
        save_name_str = './/Excel Files/' + str(exercisedate[0]) + ' ' + str(savetimestart) + ' ' + str(exercisetype) + ' Data.xlsx'
        save_name = Path(__file__).parent / save_name_str
        
    except:
        save_name = potential_save_name

    # Check if any files will be overwritten and prevent this by changing the name
    for i in excel_filelocations:
        if str(i) == str(save_name):
            save_name = potential_save_name      
            print('Issue naming file. A file already exists with the expected name. A temporary name will be used.\n')
    
    # Save the file in an excel document using the date and exercise type to create a new file each time
    try:
        wb.save(save_name) # This will overwrite any existing file in this location
    except:
        pass
    
    print('File converted: ' + str(exercisetype) + ' ' + str(exercisedate[0]) + ' ' + str(saved_data[0][0]))

    # Add the path of the file which has been converted, so it is not converted in future iterations. 
    convertedfiles_file = open(converted_list_path, 'a')
    convertedfiles_file.write(str(file) + '\n')
    convertedfiles_file.close()

print()
print()




    
    